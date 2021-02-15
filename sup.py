from openpyxl import load_workbook

class result:
    def __init__(self, name, branch):
        self.name = name
        self.branch = branch
        self.excelfiles = ['dat/B. TECH. I SEM DEC 18.xlsx', 'dat/B. TECH. II SEM JUNE 2019.xlsx',
                           'dat/B. TECH. III SEM DECEMBER 2019.xlsx', 'dat/B. TECH. IV SEM DECEMBER 2020.xlsx']
        self.letter = 'a'
        self.x = []
        self.total_marks = [0, 0]
        self.select()

    def select(self):

        for self.sem in range(0, 4):
            self.semn = load_workbook(
                self.excelfiles[self.sem], data_only=True)
            self.cS = self.semn[self.branch]

            self.search()

            if self.letter == 'a':
                print('data not matching')
                if self.sem == 2:
                    break
                continue

            self.vals()

    def search(self):   # searching name
        for row in range(7, self.cS.max_row + 1):
            for column in "DE":
                self.cell_name = "{}{}".format(column, row)
                ex = self.cS[self.cell_name].value
                if type(ex) == str:
                    ex = ex.strip()
                    ex = ex.lower()
                if ex == self.name:
                    self.letter = row
                    return 0

    def vals(self):                          # saving values for table and plot
        row = 4
        for column in range(1, self.cS.max_column + 1):
            tem = self.cS.cell(row, column).value
            if type(tem) == str:
                tem = tem.strip()
                tem = tem.lower()
                if tem == 'result':
                    column -= 1
                    row += 2
                    self.sem += 1
                    self.x.append([self.sem, "{}/{}".format(self.cS.cell(self.letter, column).value, self.cS.cell(row, column).value),
                                round(self.cS.cell(self.letter, column).value / self.cS.cell(row, column).value * 100, 4)])
                    self.total_marks[0] = self.total_marks[0] + self.cS.cell(self.letter, column).value
                    self.total_marks[1] = self.total_marks[1] + self.cS.cell(row, column).value
                    return 0

    def display(self, v):
        if v == 'check':
            return self.letter
        elif v == 't':
            self.x.append(["Total:", "{}/{}".format(self.total_marks[0], self.total_marks[1]),
                       round(self.total_marks[0]/self.total_marks[1] * 100, 4)])
            return self.x

    def clear(self):
        self.x.clear
